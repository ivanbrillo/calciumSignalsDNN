import os

import numpy as np
import openpyxl
import pandas as pd


def create_dataframe(path: str = './'):
    xlsx_files = [path + "/" + file for file in os.listdir(path) if file.endswith(".xlsx")]
    series_list = list()

    for path in xlsx_files:
        wb_obj = openpyxl.load_workbook(path)

        for page in wb_obj.sheetnames:
            sheet = wb_obj[page]

            for column in sheet.iter_cols(values_only=True):
                time_data = pd.Series(data=column)

                if time_data is None or time_data.empty or time_data.isna().sum() == time_data.size:
                    continue

                data_dict = multi_page_dict_creator(page, time_data.to_numpy())

                if not any(element is None for element in time_data):
                    series_list.append(data_dict)

    return pd.DataFrame(series_list)


def multi_page_dict_creator(page: str, time_data) -> dict:
    page_list = page.split(' ')

    if len(time_data) == 1799:
        time_data = np.append(time_data, [time_data[-1], ])

    labels = {'source': "DBD", 'stimulus': "PAW", 'date': int(''.join(filter(str.isdigit, page_list[0]))),
              'activation_time': str(int(''.join(filter(str.isdigit, page_list[1]))) * 60) + "s",
              'frequency_power': str(int(''.join(filter(str.isdigit, page_list[2])))) + "kHz",
              'time_series': time_data,
              # 'shape_ts': time_data.shape
              }

    if "RT " in page:
        labels['storage_duration'] = 0
    else:
        value = int(''.join(filter(str.isdigit, page_list[-1])))
        if "d" in page_list[-1]:
            labels['storage_duration'] = value
        elif "m" in page_list[-1]:
            labels['storage_duration'] = value * 30
        else:
            labels['storage_duration'] = value * 7

        # labels['temperature'] = 25

    return labels


# df = create_dataframe()
# pd.set_option('display.max_rows', None)  # Display all rows
# pd.set_option('display.max_columns', None)  # Display all columns
# pd.set_option('display.width', None)  # Adjust display width
# pd.set_option('display.max_colwidth', None)  # Display entire content of columns
#
# print(df)
